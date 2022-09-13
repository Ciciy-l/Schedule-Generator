import sys

import PyQt5.QtCore
from PyQt5 import QtCore, QtGui, QtWidgets
from main_ui import Ui_MainWindow

import calendar
import datetime
import json
import os
import requests
import openpyxl.styles
from openpyxl import Workbook, load_workbook, worksheet
import time


class Schedule(object):

    def __init__(self, month: int, personal_file_name: str, xlsx_file_name: str, skip):
        self.current_time = time.localtime()
        self.year = time.strftime("%Y", time.localtime())
        self.month = month
        self.data_path = os.path.abspath(rf"personal_information\{personal_file_name}.txt")
        self.output_xlsx_path = os.path.abspath(f"{xlsx_file_name}.xlsx")
        if skip == "是":
            self.skip_holidays = True
        elif skip == "否":
            self.skip_holidays = False
        self.file_exists(self.data_path)
        self.file_exists(self.output_xlsx_path)

    def file_exists(self, path):
        if not os.path.exists(path):
            print(f"文件:{path}不存在，请检查！")
            raise FileExistsError

    def read_personal_information(self):
        data_path = self.data_path
        personal_list = []
        with open(data_path, 'r+', encoding='utf-8') as f:
            for line in f.readlines():
                personal_list.append(line.strip())
        f.close()
        with open(data_path.replace("txt", "bak"), 'w', encoding='utf-8') as f:
            f.write('\n'.join(personal_list))
            f.close()
        return personal_list

    def update_personal_information(self, start_num: str, personal_list: list):
        data_path = self.data_path
        with open(data_path, 'w', encoding='utf-8') as f:
            personal_list.insert(0, start_num)
            f.write('\n'.join(personal_list))
            f.close()
        return personal_list

    def current_personal_list(self, read_list: list):
        start_cur = int(read_list[0])
        if start_cur != 1:
            first_half_list = read_list[start_cur::]
            second_half_list = read_list[1:start_cur]
            return first_half_list + second_half_list
        else:
            del read_list[0]
            return read_list

    def creation_date(self):
        year = int(self.year)
        month = self.month
        date_list = []
        for i in range(calendar.monthrange(year, month)[1] + 1)[1:]:
            str1 = f"{year}年{month}月{i}日"
            date_list.append(str1)
        return date_list

    def is_holiday(self, year, month, day):
        response = requests.get(
            f"https://api.apihubs.cn/holiday/get?field=date,week,workday,holiday_recess&year={year}&month={year}{month}&cn=1&size=31")
        day_dict = json.loads(response.text).get("data").get("list")[::-1][day - 1]
        if day_dict.get("workday_cn") == "非工作日" and day_dict.get("holiday_recess_cn") == "假期节假日":
            return [True, day_dict.get("date_cn"), day_dict.get("week_cn")]
        return [False, day_dict.get("date_cn"), day_dict.get("week_cn")]

    def output_xlsx(self, date_list, personal_list):
        output_xlsx_path = self.output_xlsx_path
        current_personal_list = self.current_personal_list(personal_list)

        # 尝试打开并保存excel文件，确定是否可写入
        wb = load_workbook(output_xlsx_path)
        wb.save(output_xlsx_path)
        wb.close()

        max_id = len(current_personal_list) - 1
        year = int(date_list[0][0:4])
        wb = load_workbook(output_xlsx_path)
        ws = wb.active
        fianl_row_num = int(date_list[-1].split("月")[-1].split("日")[0]) + 2
        name_id = -1
        for row_num, date, day in zip(range(3, 35), date_list, range(1, 32)):
            ws.cell(row_num, 1, value=date.split("年")[-1]).font = openpyxl.styles.Font(name=u'宋体', size=10,
                                                                                        bold=False, color='000000')
            weekday = datetime.date(year, self.month, int(date.split("月")[-1].split("日")[0])).weekday()
            holiday = self.is_holiday(str(self.year), str(self.month).zfill(2), day)
            if not self.skip_holidays:
                holiday[0] = False
            if not holiday[0]:
                if name_id == max_id:
                    name_id = -1
                name_id += 1
                personal = current_personal_list[name_id]
                if weekday in [0, 1, 2, 3, 4]:
                    prefix = "夜班："
                    font_bold = False
                else:
                    prefix = "全天："
                    font_bold = True
                ws.cell(row_num, 3, value=f"{prefix}{personal}").font = openpyxl.styles.Font(name=u'宋体', size=10,
                                                                                             bold=font_bold,
                                                                                             color='000000')
            else:
                ws.cell(row_num, 3, value="\\").font = openpyxl.styles.Font(name=u'宋体', size=10, bold=False,
                                                                            color='000000')

        if 33 - fianl_row_num > 0:
            [ws.cell(33 - num, 1, value="\\") for num in range(33 - fianl_row_num)]
            [ws.cell(33 - num, 3, value="\\") for num in range(33 - fianl_row_num)]

        self.update_personal_information(str(name_id + 2), current_personal_list)
        wb.save(output_xlsx_path)
        wb.close()
        print(f"{self.month}月值班表已填入excel表格!")

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.btn_click()

    def btn_click(self):
        self.ui.pushButton.clicked.connect(self.main_func)

    def main_func(self):
        self.ui.label_4.setText("正在写入Excel，请稍后……")
        radioButton_month_list = [self.ui.radioButton, self.ui.radioButton_2, self.ui.radioButton_3, self.ui.radioButton_4,
                                  self.ui.radioButton_5,
                                  self.ui.radioButton_6, self.ui.radioButton_7, self.ui.radioButton_8, self.ui.radioButton_9,
                                  self.ui.radioButton_10,
                                  self.ui.radioButton_11, self.ui.radioButton_12]
        month = [radioButton.text().split("月")[0] for radioButton in radioButton_month_list if radioButton.isChecked()][0]
        radioButton_skip_list = [self.ui.radioButton_14, self.ui.radioButton_13]
        skip_holidays = [radioButton.text() for radioButton in radioButton_skip_list if radioButton.isChecked()][0]
        file = self.ui.lineEdit.text()
        main_function = Schedule(int(month), personal_file_name="personal", xlsx_file_name=file, skip=skip_holidays)
        while True:
            try:
                main_function.output_xlsx(main_function.creation_date(), main_function.read_personal_information())
            except PermissionError:
                self.ui.label_4.setText("Excel文件已被其他应用占用！请关闭占用软件后重试…")
            else:
                self.ui.label_4.setText("排班表已填写完成!")
                break

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)  # 实例化一个应用对象
    MainWindow = MainWindow()
    MainWindow.show()  # 显示窗口
    sys.exit(app.exec_())  # 程序循环,等待安全退出