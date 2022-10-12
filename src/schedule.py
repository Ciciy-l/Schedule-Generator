# -*- coding: UTF-8 -*-
# by:Caiqiancheng
# Date:2022/9/16
import calendar
import datetime
import json
import os
import requests
import openpyxl.styles
from openpyxl import Workbook, load_workbook, worksheet
from src.common import read_config
from src.excelOperation import ExcelOperations
import time


class Schedule(object):

    def __init__(self, month: int, personal_file_name: str, xlsx_template_name: str, skip, xlsx_output_name: str):
        self.current_time = time.localtime()
        self.year = time.strftime("%Y", time.localtime())
        self.month = month
        # 工作人员信息文件路径
        self.staff_data_path = os.path.abspath(rf"personal_information\{personal_file_name}.txt")
        # 领导人员信息文件路径
        self.leader_path = os.path.abspath(r"personal_information\leader_information.txt")
        if not xlsx_template_name:
            xlsx_template_name = read_config("default").get("xlsx_template_name")
        if not xlsx_output_name:
            xlsx_output_name = read_config("default").get("output_xlsx_name")
        self.template_xlsx_path = os.path.abspath(rf"res\{xlsx_template_name}.xlsx")
        self.output_xlsx_path = os.path.abspath(f"{xlsx_output_name}.xlsx")

        if skip == "y" or skip == "是":
            self.skip_holidays = True
        elif skip == "n" or skip == "否":
            self.skip_holidays = False

        self.file_exists(self.staff_data_path)
        self.file_exists(self.template_xlsx_path)

        # 实例化 ExcelOperations 类
        self.excel_operation = ExcelOperations(self.template_xlsx_path)
        # 标签替换并保存
        self.excel_operation.replacing_labels_in_regions(mode="time")
        self.excel_operation.replacing_labels_in_regions(mode="text")
        self.excel_operation.replacing_labels_in_regions(
            input_replace_dict={"{{year}}": self.year, "{{mouth}}": self.month})
        self.excel_operation.save_wb(self.output_xlsx_path)
        self.excel_operation.close_wb()

    def file_exists(self, path):
        if not os.path.exists(path):
            print(f"文件:{path}不存在，请检查！")
            raise FileExistsError

    def read_personal_information(self, data_path="staff"):
        if data_path == "staff":
            data_path = self.staff_data_path
        elif data_path == "leader":
            data_path = self.leader_path
        personal_list = []
        with open(data_path, 'r+', encoding='utf-8') as f:
            for line in f.readlines():
                personal_list.append(line.strip())
        f.close()
        with open(data_path.replace("txt", "bak"), 'w', encoding='utf-8') as f:
            f.write('\n'.join(personal_list))
            f.close()
        return personal_list

    def update_personal_information(self, start_num: str, personal_list: list, data_path: str):
        if data_path == "staff":
            data_path = self.staff_data_path
        elif data_path == "leader":
            data_path = self.leader_path
        with open(data_path, 'w', encoding='utf-8') as f:
            personal_list.insert(0, start_num)
            f.write('\n'.join(personal_list))
            f.close()
        return personal_list

    def current_personal_list(self, read_list: list):
        start_cur = int(read_list[0])
        if start_cur != 1:
            first_half_list = read_list[start_cur::]
            second_half_list = read_list[1:start_cur]
            return first_half_list + second_half_list
        else:
            del read_list[0]
            return read_list

    def creation_date(self):
        year = int(self.year)
        month = self.month
        date_list = []
        for i in range(calendar.monthrange(year, month)[1] + 1)[1:]:
            str1 = f"{year}年{month}月{i}日"
            date_list.append(str1)
        return date_list

    def get_date_info(self, year, month):
        response = requests.get(
            f"https://api.apihubs.cn/holiday/get?field=date,week,workday,holiday_recess&year={year}&month={year}{month}&cn=1&size=31")
        date_info = dict(json.loads(response.text))
        with open(file="./res/holiday_information/{}{}.json".format(year, month), mode="w+", encoding="utf-8") as f:
            f.write(str(response.text))
        return date_info

    def is_holiday_day(self, date_info, day):
        day_dict = date_info.get("data").get("list")[::-1][day - 1]
        if day_dict.get("workday_cn") == "非工作日" and day_dict.get("holiday_recess_cn") == "假期节假日":
            return [True, day_dict.get("date_cn"), day_dict.get("week_cn")]
        return [False, day_dict.get("date_cn"), day_dict.get("week_cn")]

    def output_xlsx(self, date_list, staff_list, leader_list):
        def index_loop(rang: list, index: int):
            start_index = rang[0]
            end_index = rang[1]
            if index in range(start_index, end_index + 1):
                if index + 1 <= end_index:
                    index += 1
                else:
                    index = start_index
            return index

        output_xlsx_path = self.output_xlsx_path
        current_staff_list = self.current_personal_list(staff_list)
        current_leader_list = self.current_personal_list(leader_list)

        # 尝试打开并保存excel文件，确定是否可写入
        wb = load_workbook(output_xlsx_path)
        wb.save(output_xlsx_path)
        wb.close()

        max_id = len(current_staff_list) - 1
        year = int(date_list[0][0:4])
        wb = load_workbook(output_xlsx_path)
        ws = wb.active
        fianl_row_num = int(date_list[-1].split("月")[-1].split("日")[0]) + 2
        name_id = -1
        leader_index = 1
        merge_row_start = 3
        date_info = self.get_date_info(str(self.year), str(self.month).zfill(2))
        for row_num, date, day in zip(range(3, 35), date_list, range(1, 32)):
            ws.cell(row_num, 1, value=date.split("年")[-1]).font = openpyxl.styles.Font(name=u'宋体', size=10,
                                                                                        bold=False, color='000000')
            weekday = datetime.date(year, self.month, int(date.split("月")[-1].split("日")[0])).weekday()
            holiday = self.is_holiday_day(date_info, day)
            if not self.skip_holidays:
                holiday[0] = False
            if not holiday[0]:
                if name_id == max_id:
                    name_id = -1
                name_id += 1
                personal = current_staff_list[name_id]
                if weekday in [0, 1, 2, 3, 4]:
                    prefix = "夜班："
                    font_bold = False
                else:
                    prefix = "全天："
                    font_bold = True
                ws.cell(row_num, 3, value=f"{prefix}{personal}").font = openpyxl.styles.Font(name=u'宋体', size=10,
                                                                                             bold=font_bold,
                                                                                             color='000000')
            else:
                ws.cell(row_num, 3, value="\\").font = openpyxl.styles.Font(name=u'宋体', size=10, bold=False,
                                                                            color='000000')
            # 处理领导信息填写
            if row_num == 3:
                ws.cell(row_num, 2, value=current_leader_list[leader_index - 1])
            else:
                # if weekday != 0:
                #     ws.cell(row_num, 2, value=current_leader_list[leader_index - 1])
                if weekday == 0:
                    # 更新leader_index
                    leader_index = index_loop([1, len(current_leader_list)], leader_index)
                    # 填入领导信息
                    ws.cell(row_num, 2, value=current_leader_list[leader_index - 1])
                    # 合并之前的单元格
                    merge_row_end = row_num - 1
                    ws.merge_cells(start_row=merge_row_start, start_column=2, end_row=merge_row_end, end_column=2)
                    # 更新合并范围
                    merge_row_start = row_num
                if row_num == fianl_row_num:
                    # 最后一次合并单元格
                    ws.merge_cells(start_row=merge_row_start, start_column=2, end_row=fianl_row_num, end_column=2)

        if 33 - fianl_row_num > 0:
            [ws.cell(33 - num, 1, value="\\") for num in range(33 - fianl_row_num)]
            [ws.cell(33 - num, 2, value="\\") for num in range(33 - fianl_row_num)]
            [ws.cell(33 - num, 3, value="\\") for num in range(33 - fianl_row_num)]

        self.update_personal_information(str(name_id + 2), current_staff_list, data_path="staff")
        self.update_personal_information(str(leader_index), current_leader_list, data_path="leader")
        wb.save(output_xlsx_path)
        wb.close()
        print(f"{self.month}月值班表已填入excel表格!")
