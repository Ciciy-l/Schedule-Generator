# -*- coding: UTF-8 -*-
# by:Caiqiancheng
# Date:2022/9/16
# -*- coding: UTF-8 -*-
import os
import time

from openpyxl import load_workbook
from src.common import read_config


class ExcelOperations:

    def __init__(self, excel_path):
        self.excel_path = os.path.abspath(excel_path)
        self.wb = load_workbook(self.excel_path)

    def get_sheet_list(self, list_type="name"):
        """
        获取sheet名称或对象列表
        :param list_type:"name"返回名称列表，"obj"返回对象列表
        """
        sheet_list = []
        if list_type == "name":
            for sheet in self.wb.sheetnames:
                sheet_list.append(sheet)
        elif list_type == "obj":
            for sheet in self.wb:
                sheet_list.append(sheet)
        return sheet_list

    def get_sheet_data(self, sheet_name=None):
        """
        传入sheet名称，将每行的数据以字典的形式获取，并放在字典中返回
        :param sheet_name: sheet名称
        :return: 以[{key1:value1,key2:value1...},{key1:value2,key2:value3...}]返回sheet数据
        """
        if sheet_name:
            sheet_obj = self.wb[sheet_name]
        else:
            sheet_obj = self.wb.active
        key_tuple = sheet_obj[1]
        key_list = []
        sheet_data_list = []
        [key_list.append(key_cell.value) for key_cell in key_tuple]
        for row in sheet_obj.iter_rows(min_row=2, max_row=sheet_obj.max_row):
            value_list = []
            data_dict = {}
            [value_list.append(cell.value) for cell in row]
            [data_dict.update({key: value}) for key, value in zip(key_list, value_list)]
            sheet_data_list.append(data_dict)
        return sheet_data_list

    def get_sheet_datalines(self, sheet_name=None):
        """
        按行获取sheet数据
        :param sheet_name: sheet名称
        :return:[[line1_cell1,line1_cell2,...],[line2_cell1,line3_cell2,...],...]
        """
        if sheet_name:
            sheet_obj = self.wb[sheet_name]
        else:
            sheet_obj = self.wb.active
        data_list = []
        for row in sheet_obj.iter_rows():
            cell_list = []
            for cell in row:
                cell_list.append(cell.value)
            data_list.append(cell_list)
        return data_list

    def insert_blank_line(self, position: int, rows: int):
        """
        在指定行之前插入若干行
        :param position: 指定行号
        :param rows: 插入行数目
        """
        current_sheet = self.wb.active
        [current_sheet.insert_rows(position) for _ in range(rows)]

    def hidden_data_lines(self, position: int, rows: int, sheet_name=None):
        """
        从指定行之后隐藏若干行
        :param position: 指定行号
        :param rows: 隐藏行数目
        :param sheet_name: sheet名称
        """
        if sheet_name:
            current_sheet = self.wb[sheet_name]
        else:
            current_sheet = self.wb.active

        for row in range(position, position + rows):
            current_sheet.row_dimensions[row].hidden = 1

    def replacing_labels_in_regions(self, sheet_name=None, mode="text"):
        """
        替换excel中的标签
        :param sheet_name:
        :param mode: "text"-普通文本替换 "time"-时间戳格式化替换
        """
        if sheet_name:
            sheet_obj = self.wb[sheet_name]
        else:
            sheet_obj = self.wb.active
        replace_dict = read_config("replace_labels")
        for row in sheet_obj.iter_rows():
            for cell in row:
                if cell.value:
                    cell_text = str(cell.value)
                    for label in replace_dict.keys():
                        if label in cell_text:
                            if mode == "time":
                                new_text = cell_text.replace(label, time.strftime(replace_dict.get(label)))
                            elif mode == "text":
                                new_text = cell_text.replace(label, replace_dict.get(label))
                            else:
                                new_text = cell_text.replace(label, replace_dict.get(label))
                            cell.value = new_text

    def delete_sheet(self, sheet_name=None):
        """
        删除sheet
        :param sheet_name: 要删除的sheet名称
        """
        if sheet_name:
            del self.wb[sheet_name]
        else:
            del self.wb.active

    def save_wb(self, file_path=None):
        """
        保存修改，默认覆盖保存，也可指定保存路径
        """
        file_path = self.excel_path if not file_path else file_path
        self.wb.save(file_path)
        self.wb.close()

    def close_wb(self):
        """
        关闭工作簿
        """
        self.wb.close()

    def switch_sheet(self, sheet_name=None):
        """
        切换sheet页
        :param sheet_name: 需要切换的sheet名称
        """
        if sheet_name:
            sheet_obj = self.wb[sheet_name]
        else:
            sheet_obj = self.wb.active
        return sheet_obj
